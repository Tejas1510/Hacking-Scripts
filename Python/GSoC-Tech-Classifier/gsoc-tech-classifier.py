import requests
import openpyxl
import bs4
import os
import argparse
from fake_useragent import UserAgent

ua = UserAgent() 
header = { 
    "User-Agent": ua.random 
     } 

ap = argparse.ArgumentParser()
ap.add_argument("-t","--tech",required=True,help="Technology for which Organizations to be searched")
args = vars(ap.parse_args())

technology = args["tech"]
str(technology)

url = "https://summerofcode.withgoogle.com/archive/2020/organizations/"

res = requests.get(url)
res.raise_for_status()

soup = bs4.BeautifulSoup(res.text,'html.parser')

organizations = soup.select('h4[class="organization-card__name font-black-54"]')

all_org_Link = soup.find_all("a",class_="organization-card__link")
tech_Status = ['No']*len(organizations)
org_Tech_URL = ['none']*len(all_org_Link)

tech_index = 0

for link in all_org_Link:
    
    comp_Link = link.get('href')
    comp_url = "https://summerofcode.withgoogle.com" + comp_Link

    print(tech_index)
    print(comp_url)

    org_Tech_URL[tech_index] = comp_url
    res2 = requests.get(comp_url)
    res2.raise_for_status()

    soup2 = bs4.BeautifulSoup(res2.text,'html.parser')

    comp_Tech = soup2.find_all("li",class_="organization__tag organization__tag--technology")

    for name in comp_Tech:

        if technology in name.getText():
            tech_Status[tech_index] = 'Yes'

    tech_index = tech_index + 1

wb = openpyxl.Workbook()
sheet = wb['Sheet']

sheet.cell(row=1,column=1).value="Organization"

sheet.cell(row=1,column=2).value="Does {0} Technology Offered?".format(technology)
sheet.cell(row=1,column=3).value="Link for the Organization"

for i in range(0,len(organizations)):
    sheet.cell(row = i + 2, column = 1).value= organizations[i].getText()
    sheet.cell(row = i + 2, column = 2).value = tech_Status[i] 
    sheet.cell(row = i + 2, column = 3).value = org_Tech_URL[i]

file_name = "GSoC_Org_List"

while(os.path.isfile(file_name+'.xlsx')):
    file_name = file_name+'_' 

wb.save(file_name+".xlsx")